
from openpyxl.utils import get_column_letter


def get_cells_address(
    _type: str,
    index_row_nr: int,
    index_column_nr: int,
    max_column_nr: int,
    max_row_nr: int,
):

    cells = {
        "df": "{start_col}{start_row}:{end_col}{end_row}".format(
            start_col=get_column_letter(index_column_nr),  # Letter
            end_col=get_column_letter(max_column_nr),  # Letter
            start_row=index_row_nr + 1,  # Number
            end_row=max_row_nr,  # Number
        ),
        "info": "{start_col}{start_row}:{end_col}{end_row}".format(
            start_col=get_column_letter(index_column_nr),  # Letter
            end_col=get_column_letter(max_column_nr),  # Letter
            start_row=index_row_nr - 4,  # Number
            end_row=index_row_nr + 1,  # Number
        ),
        "AT": "{start_col}{start_row}:{end_col}{end_row}".format(
            start_col=get_column_letter(max_column_nr - 1),  # Letter
            end_col=get_column_letter(max_column_nr - 1),  # Letter
            start_row=index_row_nr + 1,  # Number
            end_row=max_row_nr + 1,  # Number
        ),
        "index": "{start_col}{start_row}:{end_col}{end_row}".format(
            start_col=get_column_letter(index_column_nr),  # Letter
            end_col=get_column_letter(index_column_nr),  # Letter
            start_row=index_row_nr + 1,  # Number
            end_row=max_row_nr,  # Number
        ),

        "sum_vert": "{start_col}{start_row}:{end_col}{end_row}".format(
            start_col=get_column_letter(max_column_nr - 2),  # Letter
            end_col=get_column_letter(max_column_nr - 2),  # Letter
            start_row=index_row_nr + 1,  # Number
            end_row=max_row_nr + 1,  # Number
        ),
        "sum_horiz": "{start_col}{start_row}:{end_col}{end_row}".format(
            start_col=get_column_letter(index_column_nr),  # Letter
            end_col=get_column_letter(max_column_nr),  # Letter
            start_row=max_row_nr + 1,  # Number
            end_row=max_row_nr + 1,  # Number
        ),
        "provinces": "{start_col}{start_row}:{end_col}{end_row}".format(
            start_col=get_column_letter(index_column_nr),  # Letter
            end_col=get_column_letter(max_column_nr),  # Letter
            start_row=index_row_nr + 1,  # Number
            end_row=index_row_nr + 1,  # Number
        ),
    }

    return cells[_type]
