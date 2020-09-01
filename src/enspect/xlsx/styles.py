from openpyxl.styles import (
    Alignment, Border, Color, Fill, Font, PatternFill, Side)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
from settings import provinces, provinces_hex

COLOR_BACKGROUND = "DEE0F2"

COLOR_INFO = "DEE0F2"
COLOR_INFO_IDX = "DEE0F2"

FONT_WHITE = "DEE0F2"
FONT_GREY = "DEE0F2"
FONT_BLUE = "000000"

BORDER_COLOR = "595959"


def style_background(ws):
    cells = "{start_col}{start_row}:{end_col}{end_row}".format(
        start_col=get_column_letter(1),  # Letter
        end_col=get_column_letter(ws.max_column + 1),  # Letter
        start_row=1,  # Number
        end_row=ws.max_row + 1,  # Number
    )
    for row in ws[cells]:
        for cell in row:
            cell.fill = PatternFill(start_color=COLOR_BACKGROUND, fill_type="solid")

    set_border(ws=ws, cell_range=cells, border_style="medium")


def style_info(ws, cell, width, height):

    cells = "{start_col}{start_row}:{end_col}{end_row}".format(
        start_col=get_column_letter(cell.column),  # Letter
        end_col=get_column_letter(cell.column + width),  # Letter
        start_row=cell.row,  # Number
        end_row=cell.row + height - 1,  # Number
    )

    for row in ws[cells]:
        set_border(
            ws=ws,
            cell_range=f"{row[0].coordinate}:{row[-1].coordinate}",
            border_style="thin",
        )
        for cell in row:
            cell.fill = PatternFill(start_color="FFFFFF", fill_type="solid")
            cell.font = Font(color="FF000000", bold=False)

    set_border(ws=ws, cell_range=cells, border_style="thin")

    return cell


def style_info_index(ws, cell, width, height):

    cells = "{start_col}{start_row}:{end_col}{end_row}".format(
        start_col=get_column_letter(cell.column),  # Letter
        end_col=get_column_letter(cell.column),  # Letter
        start_row=cell.row,  # Number
        end_row=cell.row + height - 1,  # Number
    )

    for row in ws[cells]:
        for cell in row:
            cell.fill = PatternFill(start_color=COLOR_INFO_IDX, fill_type="solid")
            cell.font = Font(color=FONT_BLUE, bold=False)
            cell.alignment = Alignment(horizontal="center")
            cell.border = Border(
                left=Side(border_style="thin", color=BORDER_COLOR),
                right=Side(border_style="thin", color=BORDER_COLOR),
                top=Side(border_style="thin", color=BORDER_COLOR),
                bottom=Side(border_style="thin", color=BORDER_COLOR),
            )

    set_border(ws=ws, cell_range=cells, border_style="thin")

    return cell


# def style_df(cell):

#     if isinstance(cell.value, float) and cell.value < 1:
#         cell.number_format = "0.##0"
#     else:
#         cell.number_format = "### ### ### ##0"

#     cell.border = Border(
#         left=Side(border_style="thin", color="000000"),
#         right=Side(border_style="thin", color="000000"),
#         top=Side(border_style="thin", color="000000"),
#         bottom=Side(border_style="thin", color="000000"),
#     )

#     cell.font = Font(color="FF000000")
#     cell.alignment = Alignment(horizontal="center")
#     return cell


def style_index(cell):
    cell.fill = PatternFill(start_color=COLOR_INFO_IDX, fill_type="solid")
    cell.number_format = "###0"
    cell.font = Font(color="FF000000", italic=True, bold=True)
    cell.alignment = Alignment(horizontal="center")
    return cell


def style_sum(cell):

    if isinstance(cell.value, float) and cell.value < 1:
        cell.number_format = "0.##0"
    else:
        cell.number_format = "### ### ### ##0"

    cell.fill = PatternFill(start_color="DEE0F2", fill_type="solid")
    cell.font = Font(color="FF000000", bold=True)
    cell.alignment = Alignment(horizontal="center")
    return cell


def style_AT(cell):

    # print()
    # print("cell-val", cell.value)

    if isinstance(cell.value, str):
        cell.font = Font(color="FFFFFF", bold=True)
        return

    if isinstance(cell.value, (float, int)):

        # print("offset up")
        # print(cell.offset(0, -1).value * 1.1)

        # if (
        #     cell.value < cell.offset(0, -1).value * 1.005
        #     and cell.value > cell.offset(0, -1).value * 0.995
        # ):

        if round(cell.value, 2) == round(cell.offset(0, -1).value, 2):
            cell.fill = PatternFill(start_color="000000", fill_type="solid")

            cell.offset(0, 1).fill = PatternFill(
                start_color="000000", fill_type="solid"
            )
            cell.font = Font(color="FFFFFF", bold=False)

        else:
            cell.fill = PatternFill(start_color="F7B2B2", fill_type="solid")
            cell.offset(0, 1).fill = PatternFill(
                start_color="F7B2B2", fill_type="solid"
            )
            cell.font = Font(color="000000", bold=False)
            # cell.offset(0, 1).value = cell.value - cell.offset(0, -1).value

        # cell.number_format = "### ### ### ##0"

    # cell.font = Font(color="000000", bold=False)
    cell.alignment = Alignment(horizontal="center")

    return cell


def style_unit(ws, index_column_nr, index_row_nr, unit):

    # Style the unit range
    ws.cell(row=index_row_nr + 1, column=index_column_nr).value = unit

    # Font
    ws.cell(row=index_row_nr + 1, column=index_column_nr).font = Font(
        color="FFFFFF", bold=True
    )

    # Fill
    ws.cell(row=index_row_nr + 1, column=index_column_nr).fill = PatternFill(
        start_color="000000", fill_type="solid"
    )


def style_provinces(cell):

    if cell.value in provinces:
        fill = PatternFill(
            fill_type="solid", fgColor=Color(rgb=provinces_hex[cell.value])
        )
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = fill

    if cell.value == "Sum-AT":
        fill = PatternFill(fill_type="solid", fgColor="FFFFFF")
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = fill

    return cell


def set_border(ws, cell_range, border_style):
    rows = list(ws[cell_range])
    side = Side(border_style=border_style, color="FF000000")

    rows = list(
        rows
    )  # we convert iterator to list for simplicity, but it's not memory efficient solution
    max_y = len(rows) - 1  # index of the last row
    for pos_y, cells in enumerate(rows):
        max_x = len(cells) - 1  # index of the last cell
        for pos_x, cell in enumerate(cells):
            border = Border(
                left=cell.border.left,
                right=cell.border.right,
                top=cell.border.top,
                bottom=cell.border.bottom,
            )
            if pos_x == 0:
                border.left = side
            if pos_x == max_x:
                border.right = side
            if pos_y == 0:
                border.top = side
            if pos_y == max_y:
                border.bottom = side

            # set new border only if it's one of the edge cells
            if pos_x == 0 or pos_x == max_x or pos_y == 0 or pos_y == max_y:
                cell.border = border


def set_column_borders(ws, index_column_nr, index_row_nr, max_column_nr, max_row_nr):
    # Set border around each column in df
    for cell in ws.iter_cols(
        min_row=index_row_nr + 1,
        max_row=index_row_nr + 1,
        min_col=index_column_nr,
        max_col=max_column_nr,
    ):

        set_border(
            ws=ws,
            cell_range="{start_col}{start_row}:{end_col}{end_row}".format(
                start_col=get_column_letter(cell[0].column),  # Letter
                end_col=get_column_letter(cell[0].column),  # Letter
                start_row=index_row_nr + 2,  # Number
                end_row=max_row_nr + 1,  # Number
            ),
            border_style="medium",
        )


def fit_column_size(ws, index_column_nr, max_column_nr):
    dim_holder = DimensionHolder(worksheet=ws)

    # Fit column size
    for col in range(ws.min_column, ws.max_column):
        dim_holder[get_column_letter(col)] = ColumnDimension(
            ws,
            min=col,
            max=col,
            auto_size=True,
            bestFit=True,
        )

    return dim_holder
