
from xlsx.styles import *
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import openpyxl as opx
from typing import Union, List
import re
from pathlib import Path
import pandas as pd
from models.data import Data


def excel_tuple(address):
    """
    Takes an Excel address and converts it to a tuple.  Used as a helper for
    looking up (row, column) to pass into sheet.cell(row, col).
    >>> assert excel_tuple('A1') == (0, 0)
    >>> assert excel_tuple('Z10') == (9, 25)
    >>> assert excel_tuple('AAA999') == (998, 702)
    :type address: basestring
    :rtype: (int, int)
    """
    address = address.upper()
    m = _excel_tuple_regex.match(address)
    if not m or not m.group("col") or not m.group("row"):
        raise ValueError("Invalid Excel address {!r}".format(address))
    col = -1
    seed = 0
    for c in reversed(m.group("col")):
        col += ord(c) - ord("A") + 26 ** seed
        seed += 1
    row = int(m.group("row")) - 1
    return row, col


_excel_tuple_regex = re.compile(r"^(?P<col>[A-Z]+)(?P<row>\d+)$")


def get_workbook(path: Union[str, Path] = None):

    book = load_workbook(Path(path))
    writer = pd.ExcelWriter(path, engine="openpyxl", mode="a")
    writer.book = book

    # wb.create_sheet("Plots", 0)

    return writer


def write_to_sheet(self, wb: opx.Workbook, scaled: str, aggregates: List = None):

    # for sheet in wb.sheetnames:
    #     if not sheet == "Plots":
    #         wb.remove(wb[sheet])

    info_row_nr = 2
    index_row_nr = info_row_nr + 4
    index_column_nr = 11

    if aggregates:
        pass
    else:
        aggregates = self.data

    for aggregate in aggregates:
        print("aggregate: ", aggregate)

        # Dont make sheets for stats
        try:
            data_items = self.data[aggregate].items()
        except:
            continue

        if aggregate in wb.sheetnames[1:]:
            wb.remove(wb[aggregate])
            # ws = wb[aggregate]
        # else:
        ws = wb.create_sheet(aggregate, 1)

        # Needed for pandas.to_excel
        writer.sheets = dict((ws.title, ws) for ws in wb.worksheets)

        # data_items = [self.data[aggregate]]
        for data_type, data in data_items:
            # print("data: ", data)

            max_column_nr, len_df = self.write_df(
                writer=writer,
                ws=ws,
                scaled=scaled,
                data=data,
                data_type=data_type,
                info_row_nr=info_row_nr,
                index_row_nr=index_row_nr,
                index_column_nr=index_column_nr,
                chart="Output",
            )

            for overlay in data.overlays:

                key = list(overlay.keys())[0]

                self.write_df(
                    writer=writer,
                    ws=ws,
                    scaled=scaled,
                    data=self.data[key],
                    data_type=data_type,
                    info_row_nr=info_row_nr,
                    index_row_nr=index_row_nr,
                    index_column_nr=max_column_nr + 2,
                    chart="Overlay",
                )

            info_row_nr += len_df + 7
            print("info_row_nr: ", info_row_nr)
            index_row_nr += len_df + 7
            print("index_row_nr: ", index_row_nr)

        for data_type, data in self.kpi[aggregate].items():
            # print("data: ", data)

            self.write_df(
                writer=writer,
                ws=ws,
                scaled=scaled,
                data=data,
                data_type=data_type,
                info_row_nr=info_row_nr,
                index_row_nr=index_row_nr,
                index_column_nr=index_column_nr,
                chart="KPI",
            )

            max_column_nr, len_df = self.write_df(
                writer=writer,
                ws=ws,
                scaled=scaled,
                data=data.numerator,
                data_type=data_type,
                info_row_nr=info_row_nr,
                index_row_nr=index_row_nr,
                index_column_nr=max_column_nr + 2,
                chart="Numerator",
            )

            max_column_nr, len_df = self.write_df(
                writer=writer,
                ws=ws,
                scaled=scaled,
                data=data.denominator,
                data_type=data_type,
                info_row_nr=info_row_nr,
                index_row_nr=index_row_nr,
                index_column_nr=max_column_nr + 2,
                chart="Denominator",
            )

            info_row_nr += len_df + 7
            print("info_row_nr: ", info_row_nr)
            index_row_nr += len_df + 7
            print("index_row_nr: ", index_row_nr)

        info_row_nr += len_df + 7
        print("info_row_nr: ", info_row_nr)
        index_row_nr += len_df + 7
        print("index_row_nr: ", index_row_nr)

    writer.save()


def write_df(
    self,
    ws: opx.worksheet,
    writer: pd.ExcelWriter,
    scaled: str,
    data: Data,
    data_type: str,
    chart: str,
    info_row_nr: int,
    index_row_nr: int,
    index_column_nr: int,
):
    # Reorder the sum vert column if "AT" has been selected
    # if "AT" in provinces:

    df = data.get_data(scaled)

    if "Sum" in df.columns:
        # Swap AT and sum columns, add AT minus Sum
        df = df.reindex(columns=list(
            df.columns[:-3]) + ["Sum", "AT", "AT-Sum"])
        df["AT-Sum"] = df["AT"] - df["Sum"]

    elif "Mean" in df.columns:
        # Swap AT and mean columns, add AT minus mean
        df = df.reindex(columns=list(
            df.columns[:-3]) + ["Mean", "AT", "AT-Mean"])
        df["AT-Mean"] = df["AT"] - df["Mean"]

    max_column_nr = index_column_nr + len(df.columns)
    max_row_nr = index_row_nr + len(df.index)

    df_info = pd.DataFrame(
        index=["Title:", "Data:", "Source:", "Created:", "Chart:"],
        data=[
            data.title,
            f"{data.aggregate} - {data_type}",
            data.file,
            strftime("%Y-%m-%d %H:%M:%S", gmtime()),
            chart,
        ],
    )

    # Write df without AT and sum vert
    df.to_excel(
        writer,
        sheet_name=ws.title,
        index=True,
        header=True,
        startcol=index_column_nr - 1,
        startrow=index_row_nr,
    )

    df_info.to_excel(
        writer,
        sheet_name=ws.title,
        index=True,
        header=False,
        startcol=index_column_nr - 1,
        startrow=info_row_nr - 1,
    )

    # Style values
    for cells, style in zip(
        [
            cells_df,
            cells_info,
            cells_index,
            cells_sum_horiz,
            cells_sum_vert,
            cells_provinces,
            cells_AT,
        ],
        [
            style_df,
            style_info,
            style_index,
            style_sum,
            style_sum,
            style_provinces,
            style_AT,
        ],
    ):
        # Style whole df table
        for row in ws[cells]:
            for cell in row:
                style(cell)

    # Unit cell style
    style_unit(ws, index_column_nr, index_row_nr, data.unit)

    # Df column borders
    set_column_borders(ws, index_column_nr, index_row_nr,
                       max_column_nr, max_row_nr)

    # Specific borders
    for cells in [
        cells_df,
        cells_info,
        cells_provinces,
        cells_sum_vert,
        cells_sum_horiz,
    ]:

        set_border(ws=ws, cell_range=cells, border_style="medium")

    # Column Width
    ws.column_dimensions = fit_column_size(
        ws, index_column_nr, max_column_nr)

    # index_column_nr = ws.max_column_nr
    # max_row_nr + 3
    # print("info_row_nr: ", info_row_nr)

    return max_column_nr, len(df.index)
