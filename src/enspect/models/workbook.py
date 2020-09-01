import datetime
import os
from itertools import chain
from pathlib import Path
from typing import Iterable, List, Union

import openpyxl
import pandas as pd
import xlwings as xw
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Color, Fill, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows, expand_levels

# from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from enspect.models.data import Data


def concat_generators(*args):
    for gen in args:
        yield from gen


IDX = pd.IndexSlice


class Workbook:
    def __init__(
        self,
        name: str,
        filename: Union[str, Path],
        sheets: List = None,
        *args,
        **kwargs,
    ):
        super().__init__(*args, **kwargs)

        if not Path.exists(filename):
            wb = openpyxl.Workbook()
            wb.save(filename)
            # Path(filename).touch()

        self.book = load_workbook(filename=filename)
        self.name = name
        self.filename = filename
        # self.index_row_nr = 2
        self.height_info = 6
        self.width_df = 0

    def launch(self):
        os.system(f"start EXCEL.EXE {self.filename}")
        return

    def add_sheets(self, sheets: List):

        for sheet in sheets:

            try:
                self.book.remove(self.book[sheet])
            except BaseException:
                pass

            if sheet not in self.book.sheetnames:
                self.book.create_sheet(title=sheet)
                self.book[sheet].next_empty_row = 1
                self.book[sheet].next_empty_col = 1
                self.book[sheet].col_start_info = 1
                self.book[sheet].coordinates = []

            try:
                self.book.remove(self.book["Sheet1"])

            except BaseException:
                pass

            try:
                self.book.remove(self.book["Tabelle1"])
            except BaseException:
                pass

    def save(self):
        self.book.save(filename=self.filename)

    def write_to_sheet(
        self,
        data: pd.DataFrame,
        sheet: Worksheet,
        # scale: str = "absolute",
        # shares: str = None,
    ):
        ws = self.book[sheet]

        if data.per_energy_aggregate:

            df = data.frame

            mask = df["ES"].values == "SUM"

            df_sums_only = df[mask]

            self.write_to_cells(ws=ws, df=df_sums_only, data=data)

            if data.show_source_values_for_energy_aggregates:

                self.update_next_empty_row(
                    ws=ws, up_shift=len(df_sums_only) + self.len_info
                )

                self.update_next_empty_col(ws=ws, right_shift=len(df.columns))

                self.write_to_cells(ws=ws, df=df, data=data)

                self.update_next_empty_col(ws=ws, left_shift=len(df.columns))

        else:

            self.write_to_cells(ws=ws, df=data.frame, data=data)

    def write_to_cells(self, ws: Worksheet, df: pd.DataFrame, data: Data):

        # Add info rows in sheet
        info = self.get_info(ws=ws, data=data)

        rows = dataframe_to_rows(df, index=False)

        # Since rows and info are generators, concat them with this util func
        rows = concat_generators(info, rows)

        for r_idx, row in enumerate(rows, ws.next_empty_row):

            for c_idx, value in enumerate(row, ws.next_empty_col):
                if value is None:
                    pass
                else:
                    ws.cell(row=r_idx, column=c_idx, value=value)

        self.update_next_empty_row(ws=ws, down_shift=self.len_info + len(df))

        return

    def get_info(self, ws: Worksheet, data: pd.DataFrame):

        info = pd.DataFrame(
            index=["Title", "Data", "Scale", "Source", "Created", "ID"],
            data=[
                None,
                data.name,
                # scale,
                None,
                data.source,
                datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                data.key
                # chart,
            ],
        )

        self.len_info = len(info) + 1

        return dataframe_to_rows(info, index=True, header=False)

    @staticmethod
    def update_next_empty_row(
        ws: Worksheet, down_shift: int = None, up_shift: int = None,
    ):

        if down_shift is not None:

            # Store next row position
            ws.next_empty_row += down_shift + 1

        if up_shift is not None:

            # Store next row position
            ws.next_empty_row -= up_shift + 1

    @staticmethod
    def update_next_empty_col(
        ws: Worksheet, left_shift: int = None, right_shift: int = None
    ):

        if right_shift is not None:

            # Store next row position
            ws.next_empty_col += right_shift + 1

        if left_shift is not None:

            # Store next row position
            ws.next_empty_col -= left_shift + 1

        # Store next col position

    @staticmethod
    def reset_empty_row(ws: Worksheet):
        ws.next_empty_row = 1

    @staticmethod
    def reset_empty_col(ws: Worksheet):
        ws.next_empty_col = 1

    # if data.per_energy_aggregate:

    #     for year, years_dict in dfs.items():

    #         for energy_aggregate, energy_aggregate_dfs in years_dict.items():

    #             for enum, df in enumerate(energy_aggregate_dfs):
    #                 name = "_".join(
    #                     [
    #                         energy_aggregate.replace(" ", "_"),
    #                         "Energie_Aggregate",
    #                         str(year),
    #                     ]
    #                 )

    #                 key = "_".join(
    #                     [
    #                         "EA",
    #                         energy_aggregate.replace(" ", "_")[:3].upper(),
    #                         str(year)[2:],
    #                     ]
    #                 )

    #                 info = list(self.get_info(ws=ws, data=data, name=name, key=key))

    #                 rows = dataframe_to_rows(df, index=False)

    #                 rows = concat_generators(info, rows)

    #                 self.write_to_cells(rows=rows, ws=ws)

    #                 if enum == 1:

    #                     self.update_next_empty_row(
    #                         ws=ws, height_df=len(df), height_info=len(info)
    #                     )

    #                     self.reset_empty_col(ws=ws)

    #                 else:
    #                     self.update_next_empty_col(ws=ws, width_df=len(df.columns))

    # else:

    #     for energy_source_name, energy_source_dict in dfs.items():

    #         for (balance_aggregate_name, df,) in energy_source_dict.items():

    #             years = list(df.index)
    #             print("years: ", years)

    #             name = "_".join(
    #                 [
    #                     energy_source_name.replace(" ", "_"),
    #                     balance_aggregate_name.replace(" ", "_"),
    #                     str(years[0])[2:],
    #                     str(years[-1])[2:],
    #                 ]
    #             )

    #             key = "_".join(
    #                 [
    #                     energy_source_name.replace(" ", "_")[:3].upper(),
    #                     balance_aggregate_name.replace(" ", "_")[:3].upper(),
    #                     str(years[0])[2:],
    #                     str(years[-1])[2:],
    #                 ]
    #             )

    #             info = list(self.get_info(ws=ws, data=data, name=name, key=key))

    #             rows = dataframe_to_rows(df, index=False)

    #             rows = concat_generators(info, rows)

    #             self.write_to_cells(rows=rows, ws=ws)

    #             # if enum == 1:

    #             self.update_next_empty_row(
    #                 ws=ws, height_df=len(df), height_info=len(info)
    #             )

    # # Write dataframe row by row
    # for r_idx, row in enumerate(rows, ws.next_empty_row):
    #     for c_idx, value in enumerate(row, ws.next_empty_col):
    #         if value is None:
    #             pass
    #         else:
    #             ws.cell(row=r_idx, column=c_idx, value=value)


# class XLSX:
#     def __init__(self, name: str, filename: Union[str, Path], sheets: List = None):
#         super().__init__()

#         self.wb = xw.Book(fullname=filename)
#         self.name = name
#         self.path = filename
#         # self.index_row_nr = 2
#         # self.height_info = 6
#         # self.width_df = 0

#     def launch(self):
#         os.system(f"start EXCEL.EXE {self.path}")
#         return

#     def add_sheets(self, sheets: List):

#         for sheet in sheets:
#             print("sheet: ", sheet)

#             for sh in self.wb.sheets:
#                 if sheet in sh.name:
#                     sh.delete()

#             # if sheet in self.wb.sheets:
#             #     print("self.wb.sheets: ", self.wb.sheets)
#             #     self.wb.sheets[sheet].delete()

#             # try:
#             #     self.book.remove(self.book[sheet])
#             # except BaseException:
#             #     pass

#             else:
#                 print("sss")
#                 self.wb.sheets.add(sheet)
#                 # self.book[sheet].next_empty_row = 1
#                 # print(
#                 #     "self.book[sheet].next_empty_row : ",
#                 #     self.book[sheet].next_empty_row,
#                 # )
#                 # self.book[sheet].next_empty_col = 1
#                 # self.book[sheet].col_start_info = 1
#                 # self.book[sheet].coordinates = []

#             # try:
#             #     self.book.remove(self.book["Sheet1"])

#             # except BaseException:
#             #     pass

#             # try:
#             #     self.book.remove(self.book["Tabelle1"])
#             # except BaseException:
#             #     pass

#     def write_to_sheet(
#         self,
#         data: pd.DataFrame,
#         sheet: Worksheet,
#         # scale: str = "absolute",
#         # shares: str = None,
#     ):
#         ws = self.wb.sheets[sheet]
#         # print("ws: ", ws.next_empty_row)

#         dfs = data.xlsx_formatted()

#         # df = df.stack(level="PROV")

#         # df.reset_index(level="YEAR", col_level=0, inplace=True)
#         # df.reset_index(col_level="PROV", inplace=True)
#         # df.reset_index(self, drop=True, inplace=True)
#         # print("df: ", df)
#         # print("data: ", data)

#         # self.update_sheet_info(ws=ws, df=df)

#         # ws.next_row_start = ws.max_row - len(df.index) - 1
#         # # ws.row_end_info = ws.min_row
#         # ws.next_col_start = ws.min_column + len(df.columns) + 2
#         # # ws.col_end_info = ws.max_column

#         self.add_data(ws=ws, dfs=dfs, data=data, unit=data.unit)

#         # self.book.save(filename="test.xlsx")

#     @staticmethod
# def update_next_empty_row(ws: Worksheet, len_df: int, len_info: int =
# None):

#         # Store next row position
#         ws.next_empty_row += len_df + len_info + 1
#         # print("ws.next_empty_row: ", ws.next_empty_row)

#         # # Store next row position
#         # ws.next_empty_col += len(df.columns) + 3
#         # print("ws.next_empty_col: ", ws.next_empty_col)

# def add_data(self, data: Data, ws: Worksheet, dfs: pd.DataFrame, unit:
# str):

#         for energy_source, aggregates in dfs.items():

#             for key, df in aggregates.items():
#                 print("df: ", df)
#                 ws.range("A1").values = df
#             #     info = list(self.get_info(ws=ws, data=data))
#             #     rows = dataframe_to_rows(df, index=False)

#             #     rows = concat_generators(info, rows)

#             #     for r_idx, row in enumerate(rows, ws.next_empty_row):

#             #         # if r_idx == ws.next_empty_row:
#             #         #     row.pop(0)
#             #         #     row.insert(0, unit)

#             #         for c_idx, value in enumerate(row, ws.next_empty_col):
#             #             if value is None:
#             #                 pass
#             #             else:
#             #                 ws.cell(row=r_idx, column=c_idx, value=value)

#             #     self.update_next_empty_row(
#             #         ws=ws, len_df=len(df), len_info=len(list(info))
#             #     )

#     @staticmethod
#     def get_info(ws: Worksheet, data: pd.DataFrame):
#         info = pd.DataFrame(
#             index=["Title", "Data", "Scale", "Source", "Created", "Chart"],
#             data=[
#                 None,
#                 data.name,
#                 # scale,
#                 None,
#                 data.source,
#                 datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
#                 ""
#                 # chart,
#             ],
#         )

#         return dataframe_to_rows(info, index=True, header=False)

#         # # Write dataframe row by row
#         # for r_idx, row in enumerate(rows, ws.next_empty_row):
#         #     for c_idx, value in enumerate(row, ws.next_empty_col):
#         #         if value is None:
#         #             pass
#         #         else:
#         #             ws.cell(row=r_idx, column=c_idx, value=value)
