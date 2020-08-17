import openpyxl as opx
from openpyxl.utils import get_column_letter

# from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from xlsx.styles import *
from pathlib import Path
import pandas as pd
from typing import Union, List
import os
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Font, Alignment, PatternFill, Fill, Color
from models.data import Data
import datetime


class xlsx:
    def __init__(self, name: str, path: Union[str, Path], sheets: List = None):
        super().__init__()

        self.book = load_workbook(filename=path)
        self.name = name
        self.path = path
        # self.index_row_nr = 2
        self.height_info = 6
        self.width_df = 0

    def launch(self):
        os.system(f"start EXCEL.EXE {self.path}")
        return

    def add_sheets(self, sheets: List):

        for sheet in sheets:

            try:
                self.book.remove(self.book[sheet])
            except BaseException:
                pass

            if sheet not in self.book.sheetnames:

                self.book.create_sheet(title=sheet)
                self.book[sheet].row_start_info = 1
                self.book[sheet].row_end_info = 1
                self.book[sheet].col_start_info = 1
                self.book[sheet].coordinates = []

            try:
                self.book.remove(self.book["Sheet1"])

            except BaseException:
                pass

            try:
                self.book.remove(self.book["Tabelle1"])
            except BaseException:
                pass

    def write(
        self,
        data: pd.DataFrame,
        sheet: Worksheet,
        scale: str = "absolute",
        shares: str = None,
    ):
        """
        write [summary]

        [extended_summary]

        Parameters
        ----------
        data : pd.DataFrame
            [description]
        sheet : Worksheet
            [description]
        scale : str
            Choose "absolute", "shares_over_rows" or "shares_over_columns"
        shares : str
            Add "shares_over_rows" or "shares_over_columns"
        """
        # self.active = self.sheetnames.index(sheet)

        ws = self.book[sheet]
        # ws.sheet_properties.Color = "1072BA"
        print("ws: ", ws)

        # # Insert empty row
        # ws.append([])

        self.insert(data=data, ws=ws, scale=scale, shares=shares)

        # rows = dataframe_to_rows(df, index=True, header=True)
        # print('rows: ', rows)

        # starts at 3 as you want to skip the first 2 rows

        # for r_idx, row in enumerate(rows, 1):

        #     # if r_idx == 1:
        #     #     ws.insert_rows(ws.max_row)

        #     ws.append(row)

        # ws.insert_rows(ws.max_row+2)

        # for c_idx, value in enumerate(row, 1):
        # print(value)
        # ws.cell(row=r_idx, column=c_idx, value=value)

        self.book.save(filename="test.xlsx")

        # wb= openpyxl.load_workbook('H:/your/dir/template.xlsx')
        # ws = wb.get_sheet_by_name('xyz')
        # rows = dataframe_to_rows(df, index=True, header=True)

        # for r_idx, row in enumerate(rows, 3):  #starts at 3 as you want to skip the first 2 rows
        #     for c_idx, value in enumerate(row, 1):
        #          ws.cell(row=r_idx, column=c_idx, value=value)

        return

    def insert(self, data: Data, ws: Worksheet, scale: str, shares: str = None):
        print("\n" + ("-" * 80))
        print("data.name: ", data.name)

        #  Main plot data
        self.add(info=True, data=data, ws=ws)

        if scale == "absolute":
            df = data.frame

        if scale == "shares_over_rows":
            df = data.shares_over_rows

        if scale == "shares_over_columns":
            df = data.shares_over_columns

        # Change columns AT and Sum, add column with difference AT-Sum
        df = self.swap_columns(df=df)

        # Create generator object from dataframe
        rows = dataframe_to_rows(df, index=True, header=True)

        # Store value for styling
        self.width_df = len(df.columns)

        for r_idx, row in enumerate(rows, 1):

            if r_idx == 1:
                row.pop(0)
                row.insert(0, data.unit)

            ws.append(row)

        ws.next_row_start = ws.max_row - len(df.index) - 1
        # ws.row_end_info = ws.min_row
        ws.next_col_start = ws.min_column + len(df.columns) + 2
        # ws.col_end_info = ws.max_column

        if shares:

            self.add(shares=shares, ws=ws, data=data, scale="Shares_" + shares)

            ws.next_col_start += len(df.columns) + 2

        if data.is_KPI:
            for data in [data.numerator, data.denominator]:
                self.add(
                    ws=ws, data=data,
                )

                ws.next_col_start += len(df.columns) + 2

        if data.has_overlay:

            for overlay in data.overlays:

                self.add(
                    ws=ws, data=overlay["data"], scale=overlay["scale"],
                )

                ws.next_col_start += len(df.columns) + 2

        ws.delete_rows(ws.row_start_info + self.height_info + 2)

        ws.row_start_info += len(data.frame) + self.height_info + 2
        ws.col_start_info = 1

        # Insert empty row
        # ws.append([])

        return

    def add(
        self,
        data: Data,
        ws: Worksheet,
        scale: str = "Absolute",
        shares: str = None,
        info: bool = False,
    ):

        if info:
            print("info.name: ", data.name)
            df = pd.DataFrame(
                index=["Title", "Data", "Scale", "Source", "Created", "Chart"],
                data=[
                    data.title,
                    data.name,
                    scale,
                    data.source,
                    datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    ""
                    # chart,
                ],
            )

            rows = dataframe_to_rows(df, index=True, header=False)

            for r_idx, row in enumerate(rows, ws.row_start_info):
                for c_idx, value in enumerate(row, ws.col_start_info):
                    if value is None:
                        pass
                    else:
                        ws.cell(row=r_idx, column=c_idx, value=value)

            ws.row_end_info += len(df) + 1

            ws.col_start_info += len(data.frame.columns) + 3
            return

        self.add(info=True, data=data, ws=ws, scale=scale)

        if scale == "shares_over_rows" or shares == "over_rows":
            df = data.shares_over_rows
            unit = "%"

        elif scale == "shares_over_columns" or shares == "over_columns":
            df = data.shares_over_columns
            unit = "%"

        else:
            df = data.frame
            unit = data.unit

        df = self.swap_columns(df=df)

        rows = dataframe_to_rows(df, index=True, header=True)

        for r_idx, row in enumerate(rows, ws.next_row_start):

            if r_idx == ws.next_row_start:
                row.pop(0)
                row.insert(0, unit)

            for c_idx, value in enumerate(row, ws.next_col_start):
                if value is None:
                    pass
                else:
                    ws.cell(row=r_idx, column=c_idx, value=value)
        return

    def style(self, ws: str):

        style_background(ws=ws)

        for col in ws.iter_cols(ws.min_column):
            for cell in col:
                if cell.value == "Title":
                    # ws.coordinates.append((cell.column, cell.row))

                    style_info(
                        ws=ws, cell=cell, width=self.width_df, height=self.height_info
                    )

                    style_info_index(
                        ws=ws, cell=cell, width=self.width_df, height=self.height_info
                    )

                    # cells_info_index = "{start_col}{start_row}:{end_col}{end_row}".format(
                    #     start_col=get_column_letter(
                    #         cell.column),  # Letter
                    #     end_col=get_column_letter(
                    #         cell.column+self.width_df),  # Letter
                    #     start_row=cell.row,  # Number
                    #     end_row=cell.row + self.height_info,  # Number
                    # )
                    # for row in ws[cells_info_index]:
                    #     for cell in row:
                    #         style_index(cell)

                    # set_border(ws=ws, cell_range=cells_info,
                    #            border_style="medium")

                    # Style info index

                    # print(cell.row, cell.column)
                    # print(ws.cell(row=cell.row, column=cell.column).value)

        # for coord in ws.coordinates:
        #     print(coord)

        return

    @staticmethod
    def swap_columns(df: pd.DataFrame):

        if "Sum" in df.columns:
            # Swap AT and sum columns, add AT minus Sum
            df = df.reindex(columns=list(df.columns[:-2]) + ["Sum", "AT", "AT-Sum"])

            df["AT-Sum"] = df["AT"] - df["Sum"]

        elif "Mean" in df.columns:
            # Swap AT and mean columns, add AT minus mean
            df = df.reindex(columns=list(df.columns[:-2]) + ["Mean", "AT", "AT-Mean"])

            df["AT-Mean"] = df["AT"] - df["Mean"]

        return df

    @staticmethod
    def search_value_in_column(ws, search_string, column="A"):
        for row in range(1, len(column) + 1):
            coordinate = "{}{}".format(column, row)
            if ws[coordinate].value == search_string:
                return column, row
        return column, None

    # @staticmethod
    # def search_value_in_col_idx(ws, search_string, col_idx=1):
    #     for row in range(1, ws.max_row + 1):
    #         if ws[row][col_idx].value == search_string:
    #             return col_idx, row
    #     return col_idx, None

    # @staticmethod
    # def search_value_in_row_index(ws, search_string, row=1):
    #     for cell in ws[row]:
    #         if cell.value == search_string:
    #             return cell.column, row
    #     return None, row
