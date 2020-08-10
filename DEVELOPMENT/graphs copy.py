from processing.data import Data
import pickle
import pandas as pd
from typing import List, Union
from pathlib import Path
from settings import file_paths
import xlwings as xw
import openpyxl as opx
from pprint import pprint

IDX = pd.IndexSlice
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import load_workbook

from openpyxl.styles import Border, Side, Font, Alignment

from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Font, NamedStyle
from openpyxl.utils import get_column_letter


index_style = NamedStyle(
    name="Index Style",
    number_format="# ###",
    font=Font(color="999999", italic=True),
    alignment=Alignment(horizontal="left"),
)

table_style = NamedStyle(
    name="Table Style",
    number_format='# ##0'
    font=Font(color="999999", italic=True),
    alignment=Alignment(horizontal="left"),
)


def format_ws(ws, cell_range):

    # applying border and alignment
    font = Font(size=9)
    align = Alignment(horizontal="left", vertical="center")
    border = Border(
        left=Side(border_style="thin", color="000000"),
        right=Side(border_style="thin", color="000000"),
        top=Side(border_style="thin", color="000000"),
        bottom=Side(border_style="thin", color="000000"),
    )

    rows = [rows for rows in ws[cell_range]]
    flattened = [item for sublist in rows for item in sublist]
    [
        (
            setattr(cell, "border", border),
            setattr(cell, "font", font),
            setattr(cell, "alignment", align),
        )
        for cell in flattened
    ]


class Graph:

    """
    One graph per aggregate, with a dataframe per energy source
    """

    def __init__(self, file_paths: List, aggregate: str):
        # Collect dataframes per energy source
        self.data = dict()  #
        self.aggregate = aggregate
        # overlay_types = ["line", "column"]
        self.overlay = dict()
        self.name = None
        self.file_paths = file_paths
        return

    def add_data(
        self,
        source: Union[str, Path],
        energy_sources: List,
        aggregate: List,
        provinces: List,
        years: List,
    ):

        if not isinstance(aggregate, list):
            aggregate = [aggregate]
        # Fetch all other kind of statistics
        if isinstance(source, Path):
            pass

        # Fetch "Energiebilanzen"
        print("provinces: ", provinces)
        if source == "EB":

            # Fetch data
            df = pickle.load(open(self.file_paths["eb"], "rb"))

            # Use the aggregate as the row index
            row_midx_addon = 5 - len(aggregate)

            # Extend the  with "Gesamt" if not other specified
            aggregate.extend(["Gesamt"] * row_midx_addon)
            # print("aggregate: ", aggregate)
            # print()

            # Aggregate name without "Gesamt"
            aggregate_name = "-".join([x for x in aggregate if x != "Gesamt"])

            # Iter over energy sources
            for energy_source in energy_sources:
                # print("energy_source: ", energy_source)

                # Slice data for all given provinces and years
                _df = df.loc[
                    IDX[tuple(aggregate)], IDX[tuple(provinces), energy_source, years]
                ].round(0)

                # Process to a df with years on index and provinces on columns
                _df = _df.to_frame().T.reset_index(drop=True).T
                _df = _df.droplevel((1), axis=0)
                _df = _df.T.stack().droplevel((0), axis=0)
                _df.index.name = ""

                # TODO: Remove after pickled index change
                # _df = _df.reindex(columns=list(df.columns[1:]) + ["AT"])
                cols = list(_df.columns)
                cols.pop(0)
                cols.append("AT")
                _df = _df[cols]

                _df["Sum"] = _df.iloc[:, :-1].sum(axis=1)

                _df = _df.T
                _df["Sum"] = _df.sum(axis=1)
                _df = _df.T

                # pprint(dir(xw.constants))

                # Create a Data object
                self.data[energy_source] = Data(
                    # absolute=pd.concat(dfs, axis=1),
                    absolute=_df,
                    # name=energy_source,
                    source=source,
                    # energy_sources=energy_sources,
                    provinces=provinces,
                    aggregate=self.aggregate,
                    years=years,
                )
            # print()
            # print("energy_source: ", energy_source)
            # print("self.data: ", self.data.keys())
            # print()

            return

    def add_overlay(self, df: pd.DataFrame):
        return

    def write_to_sheet(
        self, one_page: bool = False, clear_contents: bool = True,
    ):

        # else:
        # print("graph: ", graphs)
        # for graph in graphs:
        #     print("graph: ", graph.data)

        # wb.app.screen_updating = False

        # if clear_contents:
        #     ws.clear_contents()

        # starting_cell = ws.range("K2").adress

        book = load_workbook(file)
        writer = pd.ExcelWriter(file, engine="openpyxl")
        writer.book = book

        try:
            sheet = wb.sheets.add(self.aggregate)

        except:
            ws = wb.sheets[self.aggregate]

        if one_page:

            pass



        df_index_column = "K"
        df_index_row = 4

        for energy_source, data in self.data.items():
            # print("data: ", data)
            print("energy_source: ", energy_source)
            print()
            print("data: ", data)
            # try:
            #     ws = wb.sheets.add(data)

            # except:
            #     ws = wb.sheets[aggregate]
            file = Path("opx.xlsx")

            # df1 = pd.DataFrame({"Data": [10, 20, 30, 20, 15, 30, 45]})

            print("writer.sheets: ", writer.sheets)

            # wb = load_workbook(filename="opx.xlsx")
            # writer = pd.ExcelWriter("opx.xlsx", engine="openpyxl")

            data.absolute.to_excel(
                writer, sheet_name=self.aggregate, index=True, startcol=10, startrow=4
            )

            # for energy_source in data.energy_sources:
            #     print("energy_source: ", energy_source)
            #     # ws.autofit()

            # wb = opx.Workbook("opx.xlsx")
            # ws = wb.active

            # for row in range(1, 40):
            # ws.append(range(500, 600))

            # for r in dataframe_to_rows(data.absolute, index=True, header=True):
            #     ws.append(r)

            book.save("opx.xlsx")

            ##########################################################XLWINGS
            # Data
            # ws.range(starting_cell).value = f"Data"
            # ws.range(starting_cell).api.Font.Bold = True
            # ws.range(starting_cell.offset(0, 1)).value = energy_source

            # # Move one row down
            # next_cell = ws.range(starting_cell.offset(1, 0))

            # # Source
            # ws.range(next_cell).value = f"Source"
            # ws.range(next_cell).api.Font.Bold = True
            # ws.range(next_cell.offset(0, 1)).value = "Energiebilanzen"

            # table_cell = ws.range(starting_cell.offset(2, 0))
            # table_cell.value = data.absolute
            # table_span = table_cell.offset(1, 1).expand().address
            # print("table_span: ", table_span)
            # # ws.range(12, 12).number_format = "##.##0,,' M'"
            # ws.range(table_span).number_format = "# ##0"
            # ws.range(table_span).api.Borders.LineStyle = -4119
            # # table_cell.number_format = "£##.##0,,' M'"

            # unit_cell = ws.range(starting_cell.offset(2, 0))
            # last_row = ws.range(unit_cell.offset(1, 0).end("down").address)

            # index_column = ws.range(unit_cell.offset(1, 0), last_row)
            # index_column.color = (0, 0, 0)
            # index_column.api.Font.Color = rgb_to_int((20, 20, 255))
            # unit_cell = "TJ"

            # print("table_cell: ", table_cell)

            # # if overlay:
            # # last_column = table_cell.offset(0, 1).end("right").address
            # # starting_cell = ws.range(last_column).offset(-2, 2)
            # # print("starting_cell: ", starting_cell)

            # # else:

            # last_row = table_cell.offset(1, 0).end("down").address
            # table_index = ws.range()

            # starting_cell = ws.range(last_row).offset(2, 0)
            ######################################################################################################################################
            # print("last_row: ", last_row)
            # starting_cell = ws.range(last_row).offset(2, 0)

            # ws.range(starting_cell).value = data.absolute.astype(int)

            # print("starting_cell: ", starting_cell)

            # Source
            # ws.range("K2").value = f'Quelle: {biv_gesamt["source"]}'
            # next_row_number = ws.range("K1").end('down').row

            # ws.range("K4").options(expand="table").value = (
            #     biv_gesamt[unit].iloc[:, :] / biv_gesamt[unit].loc[2000, :]
            # )
            # ws.range("K4").value = unit

            # # next_column_number = ws.range("K5").end('right').get_address(0, 0)[0]
            # next_row_name_2 = "W4"

        # for u in energy_units:

        #     ws.range(next_row_name_2).options(
        #         expand='table').value = biv_gesamt[u].iloc[-1, :].to_frame().T

        #     ws.range(next_row_name_2).value = u

        #     next_row_number_2 = ws.range(next_row_name_2).end('down').row + 2
        #     next_row_name_2 = "W" + str(next_row_number_2)

        # next_row_name = "K" + str(next_row_number)

        # ws.range(next_row_name).value = unit
        # next_row = ws.range("K({str(next_row+1)})").end('down').row

        # next_row = ws.range(next_row_name).end('down').row
        # last_column = ws.range("K3").end('right').get_address(0, 0)[0]
        # print("The last row is {row}.".format(row=next_row))
        # print("The last col is {col}.".format(col=last_column))

        # sheet.range().expand().value = data
        # wb.app.screen_updating = True

        return

    # def compute_KPI():
    #     return
